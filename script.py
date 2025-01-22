import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
import requests
import pandas as pd
from pathlib import Path
import docx2txt
import fitz
import os
from datetime import datetime
import threading
import logging
from logging.handlers import RotatingFileHandler
import webbrowser
from PIL import Image, ImageTk
import json
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class KentekenApp:
    def __init__(self, root):
        self.root = root
        self.root.title("RDW Kenteken Checker")
        self.root.geometry("650x500")  # Increased initial window size
        self.root.minsize(650, 500) # Increased minimum window size
        self.root.resizable(True, True)

        # **Dark Mode Colors - Adjusted Button Foreground**
        self.dark_mode_bg = '#2e2e2e'  # Dark background color
        self.dark_mode_fg = '#f0f0f0'  # Light foreground color
        self.dark_mode_button_bg = '#4d4d4d' # Darker button background
        self.dark_mode_button_fg = '#0e0c0c' # **Explicitly set to white for buttons**
        self.dark_mode_accent = '#00a0dc' # Accent color (e.g., for progress bar)
        self.dark_mode_hyperlink = '#70a1ff' # Light blue for hyperlink

        # Configureer logging
        self.setup_logging()

        # Laad configuratie
        self.load_config()

        # API endpoint
        self.rdw_endpoint = "https://opendata.rdw.nl/resource/8ys7-d773.json"

        # Status variabelen
        self.processing = False
        self.current_file = None

        # Setup GUI
        self.setup_gui()

    def setup_logging(self):
        """Configureer logging."""
        log_dir = 'logs'
        os.makedirs(log_dir, exist_ok=True)

        log_file = os.path.join(log_dir, 'kenteken_checker.log')
        handler = RotatingFileHandler(
            log_file,
            maxBytes=1048576,
            backupCount=3
        )

        logging.basicConfig(
            handlers=[handler],
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )

    def setup_gui(self):
        """Setup de GUI elementen."""
        # Root window dark mode background
        self.root.configure(bg=self.dark_mode_bg)

        # Style for buttons
        button_style = ttk.Style()
        button_style.configure('TButton',
                                background=self.dark_mode_button_bg,
                                foreground=self.dark_mode_button_fg, # Use the defined button foreground color
                                relief='raised',
                                padding=8,
                                font=('Calibri', 10, 'bold'))
        button_style.map('TButton',
                          background=[('active', '#666666'), ('disabled', self.dark_mode_button_bg)]) # Hover effect

        # Style for labels
        label_style = ttk.Style()
        label_style.configure('TLabel',
                                foreground=self.dark_mode_fg,
                                background=self.dark_mode_bg,
                                font=('Calibri', 11))

        # Style for LabelFrames
        labelframe_style = ttk.Style()
        labelframe_style.configure('TLabelframe',
                                   background=self.dark_mode_bg,
                                   foreground=self.dark_mode_fg,
                                   borderwidth=2,
                                   relief='solid',
                                   font=('Calibri', 12, 'bold'))
        labelframe_style.configure('TLabelframe.Label',
                                   background=self.dark_mode_bg,
                                   foreground=self.dark_mode_fg,
                                   font=('Calibri', 12, 'bold'))


        # Style for Progressbar
        progressbar_style = ttk.Style()
        progressbar_style.configure('TProgressbar',
                                    background=self.dark_mode_accent,
                                    troughcolor=self.dark_mode_bg,
                                    borderwidth=2,
                                    relief='solid')

        # Style for Frames
        frame_style = ttk.Style()
        frame_style.configure('TFrame', background=self.dark_mode_bg)


        # Hoofdframe
        main_frame = ttk.Frame(self.root, padding="20", style='TFrame') # Apply frame style
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # Welkomstekst
        welcome_text = """
        Welkom bij de RDW Kenteken Checker!

        Deze tool kan kentekens uit PDF en Word documenten halen en opzoeken in de RDW database.

        Gebruik:
        1. Klik op 'Output map kiezen' om te bepalen waar de resultaten worden opgeslagen
        2. Klik op 'Bestand kiezen' en selecteer een PDF of Word document
        3. De tool zoekt automatisch alle kentekens en maakt een Excel bestand
        """
        welcome_label = ttk.Label(main_frame, text=welcome_text, justify=tk.LEFT, style='TLabel')
        welcome_label.grid(row=0, column=0, pady=(0, 20), sticky=tk.W)

        # Bestandsselectie frame
        select_frame = ttk.LabelFrame(main_frame, text="Documentverwerking", padding="10", style='TLabelframe') # Apply labelframe style
        select_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Output map kiezen knop
        output_button = ttk.Button(
            select_frame,
            text="Opslag Map Kiezen",
            command=self.choose_output_dir,
            style='TButton' # Apply button style
        )
        output_button.grid(row=0, column=0, padx=5, pady=5)

        # Bestand kiezen knop
        self.select_button = ttk.Button(
            select_frame,
            text="Bestand Selecteren",
            command=self.choose_file,
            style='TButton' # Apply button style
        )
        self.select_button.grid(row=1, column=0, padx=5, pady=5)

        # Status label
        self.status_var = tk.StringVar(value=f"Output map: {self.output_dir}")
        self.status_label = ttk.Label(select_frame, textvariable=self.status_var, style='TLabel')
        self.status_label.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)

        # Voortgangsbalk
        self.progress = ttk.Progressbar(select_frame, mode='indeterminate', style='TProgressbar')
        self.progress.grid(row=3, column=0, sticky=(tk.W, tk.E), padx=5, pady=5)
        self.progress.grid_remove()  # Verberg initieel

        # Help knop
        help_button = ttk.Button(
            select_frame,
            text="Help",
            command=self.show_help,
            style='TButton' # Apply button style
        )
        help_button.grid(row=4, column=0, padx=5, pady=5)

        # **Attribution Label Frame**
        attribution_frame = ttk.Frame(main_frame, style='TFrame') # Frame for attribution
        attribution_frame.grid(row=2, column=0, pady=(20, 0), sticky=(tk.W, tk.E, tk.S)) # Place below select_frame

        attribution_text = "Gemaakt door Roni Alarashye"
        attribution_label = ttk.Label(attribution_frame, text=attribution_text, style='TLabel', cursor="hand2") # Hand cursor for link
        attribution_label.grid(row=0, column=0, sticky=tk.W)
        attribution_label.configure(foreground=self.dark_mode_hyperlink) # Light blue hyperlink color


        linkedin_url = "https://www.linkedin.com/in/roni-a-88131a208/"

        def open_linkedin(event):
            webbrowser.open_new(linkedin_url)

        attribution_label.bind("<Button-1>", open_linkedin) # Bind click event to open LinkedIn


    def show_help(self):
        """Toon help informatie."""
        help_text = """
        RDW Kenteken Checker - Hulp

        Deze tool zoekt kentekens in documenten en haalt informatie op van de RDW.

        Ondersteunde bestanden:
        - PDF documenten (.pdf)
        - Word documenten (.docx)

        De kentekens moeten in het format staan:
        Kenteken: XX999X

        Het resultaat wordt opgeslagen in een Excel bestand in de gekozen map.

        Voor vragen of problemen, check de log bestanden in de 'logs' map.
        """
        messagebox.showinfo("Help", help_text)

    def choose_output_dir(self):
        """Laat gebruiker output map kiezen."""
        dir_path = filedialog.askdirectory(
            title="Kies map voor resultaten",
            initialdir=self.output_dir
        )
        if dir_path:
            self.output_dir = dir_path
            self.save_config()
            self.status_var.set(f"Output map: {dir_path}")
            messagebox.showinfo("Succes", f"Output map ingesteld op:\n{dir_path}")

    def save_config(self):
        """Sla configuratie op."""
        config_file = os.path.join(os.path.expanduser('~'), '.rdw_kenteken_config')
        with open(config_file, 'w') as f:
            json.dump({'output_dir': self.output_dir}, f)

    def load_config(self):
        """Laad configuratie."""
        config_file = os.path.join(os.path.expanduser('~'), '.rdw_kenteken_config')
        try:
            if os.path.exists(config_file):
                with open(config_file) as f:
                    config = json.load(f)
                    self.output_dir = config.get('output_dir')
            else:
                self.output_dir = os.path.expanduser('~\\Documents\\RDW Kenteken Checker')
        except:
            self.output_dir = os.path.expanduser('~\\Documents\\RDW Kenteken Checker')

    def choose_file(self):
        """Open bestandskiezer."""
        if self.processing:
            messagebox.showwarning("Bezig", "Er wordt al een bestand verwerkt")
            return

        file_path = filedialog.askopenfilename(
            title="Kies een document",
            filetypes=[
                ("Documenten", "*.pdf;*.docx"),
                ("PDF bestanden", "*.pdf"),
                ("Word documenten", "*.docx")
            ]
        )

        if file_path:
            self.process_file(file_path)

    def process_file(self, file_path):
        """Start bestandsverwerking."""
        self.processing = True
        self.select_button.config(state='disabled')
        self.progress.grid()
        self.progress.start()
        self.status_var.set("Bestand verwerken...")

        # Start verwerking in nieuwe thread
        thread = threading.Thread(target=self.process_file_thread, args=(file_path,))
        thread.daemon = True
        thread.start()

    def extract_text(self, file_path):
        """Haal tekst uit document."""
        suffix = Path(file_path).suffix.lower()
        try:
            if suffix == '.pdf':
                pdf_document = fitz.open(file_path)
                text = ""
                for page in pdf_document:
                    text += page.get_text()
                pdf_document.close()
                return text
            elif suffix == '.docx':
                return docx2txt.process(file_path)
            else:
                raise ValueError(f"Niet-ondersteund bestandsformaat: {suffix}")
        except Exception as e:
            raise Exception(f"Fout bij lezen bestand: {str(e)}")

    def extract_kentekens(self, text):
        """Zoek kentekens in tekst."""
        pattern = r'Kenteken:\s*([A-Z0-9]+)'
        matches = re.finditer(pattern, text, re.IGNORECASE)
        kentekens = []
        for match in matches:
            kenteken = match.group(1).strip()
            kenteken = re.sub(r'[-\s]', '', kenteken)
            kentekens.append(kenteken)
        return kentekens
    def get_brandstof_type(self, kenteken):
        """Haal brandstoftype op van RDW API."""
        try:
            params = {'kenteken': kenteken}
            response = requests.get(self.rdw_endpoint, params=params)
            if response.status_code == 200 and response.json():
                data = response.json()[0]
                return data.get('brandstof_omschrijving', 'Onbekend')
            return 'Niet gevonden'
        except Exception as e:
            return f'Fout: {str(e)}'

    def create_excel(self, data):
        """Maak een mooi opgemaakt Excel bestand."""
        df = pd.DataFrame(data)

        # Maak output map als die niet bestaat
        os.makedirs(self.output_dir, exist_ok=True)

        # Genereer bestandsnaam
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(self.output_dir, f'Kenteken_Resultaten_{timestamp}.xlsx')

        # Maak Excel writer met openpyxl engine
        writer = pd.ExcelWriter(excel_file, engine='openpyxl')
        df.to_excel(writer, sheet_name='Kentekens', index=False, startrow=2)

        # Haal workbook en worksheet op
        workbook = writer.book
        worksheet = writer.sheets['Kentekens']

        # Stel kolombreedtes in
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 30

        # Definieer stijlen
        title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')

        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')

        data_font = Font(name='Calibri', size=11)
        alternate_fill = PatternFill(start_color='E9EFF7', end_color='E9EFF7', fill_type='solid')

        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Voeg titel toe
        title_cell = worksheet.cell(row=1, column=1, value='RDW Kenteken Rapport')
        worksheet.merge_cells('A1:B1')
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Voeg datum toe
        date_cell = worksheet.cell(row=2, column=1,
                                 value=f'Gegenereerd op: {datetime.now().strftime("%d-%m-%Y %H:%M")}')
        worksheet.merge_cells('A2:B2')
        date_cell.font = Font(italic=True)
        date_cell.alignment = Alignment(horizontal='center')

        # Pas stijlen toe op headers (rij 3)
        for col in ['A', 'B']:
            cell = worksheet[f'{col}3']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Pas stijlen toe op data
        data_rows = worksheet.max_row
        for row in range(4, data_rows + 1):
            # Alternate row colors
            row_fill = alternate_fill if row % 2 == 0 else PatternFill()

            for col in ['A', 'B']:
                cell = worksheet[f'{col}{row}']
                cell.font = data_font
                cell.fill = row_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Voeg speciale formattering toe voor brandstoftypes
                if col == 'B':  # Brandstoftype kolom
                    value = cell.value
                    if value:
                        if 'Elektrisch' in value:
                            cell.font = Font(color='1F7A1F')  # Groen voor elektrisch
                        elif 'Benzine' in value:
                            cell.font = Font(color='000000')  # Zwart voor benzine
                        elif 'Diesel' in value:
                            cell.font = Font(color='8B0000')  # Donkerrood voor diesel

        # Auto-filter toevoegen
        worksheet.auto_filter.ref = f"A3:B{data_rows}"

        # Bevries de bovenste rijen
        worksheet.freeze_panes = 'A4'

        writer.close()
        return excel_file

    def process_file_thread(self, file_path):
        """Verwerk bestand in aparte thread."""
        try:
            self.status_var.set("Bestand lezen...")
            text = self.extract_text(file_path)

            # **Nieuwe code om de uitgehaalde tekst op te slaan:**
            output_dir = self.output_dir  # Gebruik de output directory van de app
            os.makedirs(output_dir, exist_ok=True) # Zorg dat de map bestaat
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            text_file_path = os.path.join(output_dir, f'Extracted_Text_{timestamp}.txt')
            with open(text_file_path, 'w', encoding='utf-8') as f: # Opslaan als UTF-8 om speciale tekens goed te bewaren
                f.write(text)
            self.status_var.set(f"Tekst opgeslagen in: {text_file_path}")
            # **Einde nieuwe code**

            self.status_var.set("Kentekens zoeken...")
            kentekens = self.extract_kentekens(text)

            if not kentekens:
                raise Exception("Geen kentekens gevonden in het document")

            self.status_var.set("RDW data ophalen...")
            results = []
            for kenteken in kentekens:
                brandstof = self.get_brandstof_type(kenteken)
                results.append({
                    'Kenteken': kenteken,
                    'Brandstoftype': brandstof
                })

            self.status_var.set("Excel bestand maken...")
            excel_file = self.create_excel(results)

            self.root.after(0, lambda: messagebox.showinfo("Succes",
                f"Verwerking voltooid!\n\nResultaten opgeslagen in:\n{excel_file}"))

        except Exception as e:
            self.root.after(0, lambda error_message=str(e): messagebox.showerror("Fout", error_message))
            logging.error(f"Fout bij verwerken bestand: {str(e)}")

        finally:
            self.processing = False
            self.select_button.config(state='normal')
            self.progress.stop()
            self.progress.grid_remove()
            self.status_var.set(f"Output map: {self.output_dir}")


def main():
    try:
        root = tk.Tk()

        # Windows schaling fix
        if os.name == 'nt':
            try:
                from ctypes import windll
                windll.shcore.SetProcessDpiAwareness(1)
            except:
                pass

        app = KentekenApp(root)

        # Centreer het venster
        root.update_idletasks()
        width = root.winfo_width()
        height = root.winfo_height()
        x = (root.winfo_screenwidth() // 2) - (width // 2)
        y = (root.winfo_screenheight() // 2) - (height // 2)
        root.geometry(f'{width}x{height}+{x}+{y}')

        root.mainloop()

    except Exception as e:
        logging.critical(f"Fatale fout: {str(e)}")
        messagebox.showerror("Fatale Fout",
            "Er is een onverwachte fout opgetreden. Check de logs voor meer informatie.")


if __name__ == "__main__":
    main()